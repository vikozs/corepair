"""Reporting.

Two audiences, two formats:

* **Markdown** for the platform team and the pull request.
* **XLSX** for whoever signs the purchase order. That sheet is built with *live
  formulas*, not pasted results, so the reader can change an assumption cell and
  watch the answer move. A spreadsheet of frozen numbers invites the question
  "where did this come from?"; one that recalculates answers it.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .licensing import ClusterCount, DemandTotals, Sizing
from .models import SubscriptionModel, TermOption
from .scenarios import Attribution, ScenarioResult

FONT = "Arial"
H = Font(name=FONT, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="1F4E79")
B = Font(name=FONT, bold=True)
N = Font(name=FONT)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")


# ------------------------------------------------------------------ markdown


def markdown_summary(count: ClusterCount | None, tot: DemandTotals | None,
                     sizing: Sizing | None, currency: str = "EUR") -> str:
    out: list[str] = ["# corepair report", ""]

    if count:
        out += [
            "## Current subscription footprint",
            "",
            f"- Subscribed nodes: **{len(count.subscribed_nodes)}** of {len(count.assessments)}",
            f"- Subscribed vCPU: **{count.subscribed_vcpu}**",
            f"- Core-pairs required: **{count.core_pairs}**",
            f"- Paid-for but unusable vCPU from per-node rounding: **{count.rounding_waste_vcpu}**",
            "",
        ]
        bad = count.unlabelled_infra_candidates()
        if bad:
            out += [
                "> **Finding.** The following nodes are labelled `infra` but carry no infra "
                "taint, so application pods can still land on them and the exemption does "
                "not apply. Tainting them removes them from the subscription count:",
                "",
            ]
            out += [f"> - `{a.node.name}` ({a.core_pairs} core-pairs)" for a in bad]
            out += [""]

        out += ["| Node | Role | vCPU | Core-pairs | Basis |", "|---|---|---|---|---|"]
        for a in count.assessments:
            out.append(
                f"| `{a.node.name}` | {a.node.role} | {a.node.vcpu} | "
                f"{a.core_pairs if a.subscribed else '-'} | {a.reason} |"
            )
        out.append("")

    if tot:
        out += [
            "## Requested versus actually used",
            "",
            f"- Workloads analysed: **{tot.workloads}**",
            f"- CPU requested: **{tot.cpu_request_vcpu:.1f} vCPU**",
            f"- CPU used (p95): **{tot.cpu_usage_vcpu:.1f} vCPU** "
            f"(overcommit **{tot.cpu_overcommit_ratio:.1f}x**)",
            f"- Memory requested: **{tot.mem_request_gib:.0f} GiB**",
            f"- Memory used (p95): **{tot.mem_usage_gib:.0f} GiB** "
            f"(overcommit **{tot.mem_overcommit_ratio:.1f}x**)",
            "",
            "Requests, not usage, decide what the scheduler can fit - and therefore how "
            "many nodes exist to be licensed. Where the two diverge this far, the estate "
            "is paying for reservations nobody consumes.",
            "",
        ]

    if sizing:
        out += [
            "## Sizing",
            "",
            f"- Node shape: **{sizing.shape.vcpu} vCPU / {sizing.shape.memory_gib:.0f} GiB**",
            f"- Nodes needed for CPU: **{sizing.nodes_for_cpu}**",
            f"- Nodes needed for memory: **{sizing.nodes_for_memory}**",
            f"- Binding resource: **{sizing.binding_resource.upper()}**",
            f"- Nodes (incl. {sizing.ha_spare_nodes} HA spare): **{sizing.nodes}**",
            "",
        ]
        if sizing.binding_resource == "memory":
            out += [
                "> **Finding.** Memory saturates before CPU on this shape. Adding vCPU to "
                "the node template would raise the licence count without relieving the "
                "actual constraint; adding RAM raises capacity without raising it.",
                "",
            ]
    return "\n".join(out)


def markdown_scenarios(results: list[ScenarioResult], attribution: Attribution | None,
                       currency: str, discount_rate_pct: float) -> str:
    out = ["## Cost scenarios", "", f"| Scenario | Peak units | Total ({currency}) | "
           f"NPV @ {discount_rate_pct:.0f}% ({currency}) |", "|---|---|---|---|"]
    for r in results:
        out.append(f"| {r.name} | {r.peak_units} | {r.total:,.0f} | "
                   f"{r.npv(discount_rate_pct):,.0f} |")
    out.append("")

    if attribution:
        a = attribution
        out += [
            "### Where the multi-year saving actually comes from",
            "",
            f"| Component | {currency} | Share |",
            "|---|---|---|",
            f"| Term / discount differential | {a.term_effect:,.0f} | "
            f"{100 * a.term_effect / a.total_saving if a.total_saving else 0:.0f}% |",
            f"| Avoided list-price growth (assumption) | {a.inflation_effect:,.0f} | "
            f"{100 * a.inflation_effect / a.total_saving if a.total_saving else 0:.0f}% |",
            f"| Avoided discount decay at renewal (assumption) | "
            f"{a.discount_decay_effect:,.0f} | "
            f"{100 * a.discount_decay_effect / a.total_saving if a.total_saving else 0:.0f}% |",
            f"| **Headline saving** | **{a.total_saving:,.0f}** | 100% |",
            "",
            f"**{a.assumption_driven_pct:.0f}% of the headline rests on forecasts**, not on "
            "the contract itself. Those forecasts are negotiable: a written commitment to "
            "hold the renewal discount converts an assumption into a term, and if the vendor "
            "will not commit to it, the case for the longer lock is weaker than the headline "
            "suggests.",
            "",
        ]
    return "\n".join(out)


# ---------------------------------------------------------------------- xlsx


def write_xlsx(path: str, *, currency: str, model: SubscriptionModel,
               short: TermOption, long: TermOption, units: int, horizon: int,
               discount_rate_pct: float, sizing: Sizing | None = None,
               count: ClusterCount | None = None) -> None:
    """Write a workbook whose scenario numbers are live formulas.

    Editable input cells are shaded. Everything else derives from them, so the
    reader can challenge an assumption directly instead of asking for a rerun.
    """
    wb = Workbook()

    # ---- Assumptions
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "corepair - inputs"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = "Shaded cells are inputs. Every other sheet recalculates from them."
    ws["A2"].font = Font(name=FONT, italic=True)

    rows = [
        ("Currency", currency, "From your quote"),
        ("Subscription model", model.name, model.notes or ""),
        ("SKU", model.sku, "As quoted"),
        ("List price per unit per year", model.list_price, "USER INPUT - from your own quote"),
        ("vCPU covered per unit", model.vcpu_per_unit, "Virtual core-pair = 2 cores = 4 vCPU"),
        ("Units purchased", units, "From the sizing sheet, or overridden"),
        ("Horizon (years)", horizon, "Comparison window"),
        ("NPV discount rate %", discount_rate_pct, "Your cost of capital"),
        (f"{short.name()} discount %", short.discount_pct, "USER INPUT - negotiated"),
        (f"{long.name()} discount %", long.discount_pct, "USER INPUT - negotiated"),
        ("Assumed annual list growth %", short.annual_list_growth_pct,
         "ASSUMPTION - challenge this"),
        ("Assumed renewal discount %",
         short.renewal_discount_pct if short.renewal_discount_pct is not None else short.discount_pct,
         "ASSUMPTION - ask for this in writing"),
    ]
    ws.append([])
    ws.append(["Item", "Value", "Source / note"])
    for c in ws[4]:
        c.font, c.fill = H, HFILL
    for label, value, note in rows:
        ws.append([label, value, note])
    for r in range(5, 5 + len(rows)):
        ws.cell(r, 1).font = B
        ws.cell(r, 2).font = N
        ws.cell(r, 3).font = Font(name=FONT, italic=True, size=9)
        if "INPUT" in str(ws.cell(r, 3).value) or "ASSUMPTION" in str(ws.cell(r, 3).value):
            ws.cell(r, 2).fill = INPUT_FILL

    named = {label: f"Assumptions!$B${5 + i}" for i, (label, _, _) in enumerate(rows)}
    price = named["List price per unit per year"]
    units_ref = named["Units purchased"]
    rate = named["NPV discount rate %"]
    d_short = named[f"{short.name()} discount %"]
    d_long = named[f"{long.name()} discount %"]
    growth = named["Assumed annual list growth %"]
    d_renew = named["Assumed renewal discount %"]

    # ---- Scenarios
    sc = wb.create_sheet("Scenarios")
    sc["A1"] = f"Cost over {horizon} years"
    sc["A1"].font = Font(name=FONT, bold=True, size=14)
    sc.append([])
    header = ["Year", f"{short.name()} unit price", f"{short.name()} cost",
              f"{long.name()} unit price", f"{long.name()} cost", "Discount factor"]
    sc.append(header)
    for c in sc[3]:
        c.font, c.fill = H, HFILL

    for y in range(1, horizon + 1):
        r = 3 + y
        # Short term: re-prices every year after the first, at grown list and
        # the (assumed) renewal discount.
        if y == 1:
            short_price = f"={price}*(1-{d_short}/100)"
        else:
            short_price = f"={price}*(1+{growth}/100)^{y - 1}*(1-{d_renew}/100)"
        # Long term: price held for the length of the commitment.
        if y <= long.years:
            long_price = f"={price}*(1-{d_long}/100)"
        else:
            long_price = (f"={price}*(1+{growth}/100)^{y - 1}*(1-{d_renew}/100)")
        sc.cell(r, 1, y).font = N
        sc.cell(r, 2, short_price).font = N
        sc.cell(r, 3, f"=B{r}*{units_ref}").font = N
        sc.cell(r, 4, long_price).font = N
        sc.cell(r, 5, f"=D{r}*{units_ref}").font = N
        sc.cell(r, 6, f"=1/(1+{rate}/100)^({y}-1)").font = N

    tr = 3 + horizon + 1
    sc.cell(tr, 1, "Total").font = B
    sc.cell(tr, 3, f"=SUM(C4:C{3 + horizon})").font = B
    sc.cell(tr, 5, f"=SUM(E4:E{3 + horizon})").font = B
    sc.cell(tr + 1, 1, "NPV").font = B
    sc.cell(tr + 1, 3, f"=SUMPRODUCT(C4:C{3 + horizon},F4:F{3 + horizon})").font = B
    sc.cell(tr + 1, 5, f"=SUMPRODUCT(E4:E{3 + horizon},F4:F{3 + horizon})").font = B
    sc.cell(tr + 2, 1, f"Saving of {long.name()} over {short.name()}").font = B
    sc.cell(tr + 2, 3, f"=C{tr}-E{tr}").font = B

    sc.cell(tr + 4, 1, "Attribution of that saving").font = B
    sc.cell(tr + 5, 1, "Term / discount differential only").font = N
    sc.cell(tr + 5, 3,
            f"={units_ref}*{horizon}*{price}*({d_long}-{d_short})/100").font = N
    sc.cell(tr + 6, 1, "Remainder: driven by the growth and renewal assumptions").font = N
    sc.cell(tr + 6, 3, f"=C{tr + 2}-C{tr + 5}").font = N
    sc.cell(tr + 7, 1, "Share of saving that is assumption-driven").font = B
    sc.cell(tr + 7, 3, f"=IF(C{tr + 2}=0,0,C{tr + 6}/C{tr + 2})").number_format = "0%"

    # ---- Sizing
    if sizing:
        sz = wb.create_sheet("Sizing")
        sz["A1"] = "Node count is driven by whichever resource saturates first"
        sz["A1"].font = Font(name=FONT, bold=True, size=14)
        sz.append([])
        sz.append(["Item", "Value"])
        for c in sz[3]:
            c.font, c.fill = H, HFILL
        for k, v in [
            ("Demand, CPU (vCPU)", round(sizing.cpu_vcpu, 1)),
            ("Demand, memory (GiB)", round(sizing.memory_gib, 1)),
            ("Node shape vCPU", sizing.shape.vcpu),
            ("Node shape memory (GiB)", sizing.shape.memory_gib),
            ("Headroom %", sizing.headroom_pct),
            ("Nodes needed for CPU", sizing.nodes_for_cpu),
            ("Nodes needed for memory", sizing.nodes_for_memory),
            ("HA spare nodes", sizing.ha_spare_nodes),
            ("Nodes total", sizing.nodes),
            ("Binding resource", sizing.binding_resource),
        ]:
            sz.append([k, v])
        for r in range(4, 14):
            sz.cell(r, 1).font = B
            sz.cell(r, 2).font = N

    # ---- Nodes
    if count:
        nd = wb.create_sheet("Nodes")
        nd.append(["Node", "Role", "vCPU", "Subscribed", "Core-pairs", "Basis"])
        for c in nd[1]:
            c.font, c.fill = H, HFILL
        for a in count.assessments:
            nd.append([a.node.name, a.node.role, a.node.vcpu,
                       "yes" if a.subscribed else "no", a.core_pairs, a.reason])
        for row in nd.iter_rows(min_row=2):
            for c in row:
                c.font = N
        r = nd.max_row + 2
        nd.cell(r, 1, "Total core-pairs").font = B
        nd.cell(r, 5, f"=SUM(E2:E{nd.max_row - 1})").font = B

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            width = max((len(str(sheet.cell(r, col).value or "")) for r in
                         range(1, sheet.max_row + 1)), default=10)
            sheet.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 12), 60)
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for c in row:
                if c.alignment.wrap_text is None:
                    c.alignment = Alignment(vertical="center")

    wb.save(path)
